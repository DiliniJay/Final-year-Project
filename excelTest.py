import openpyxl
import array as arr
import openpyxl
def writeCell (a):
    filename = 'example.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = a[0]
    ws.cell(row=row, column=2).value = a[1]
    ws.cell(row=row, column=3).value = a[2]
    ws.cell(row=row, column=4).value = a[3]
    ws.cell(row=row, column=5).value = a[4]
    ws.cell(row=row, column=6).value = a[5]
    ws.cell(row=row, column=7).value = a[6]
    ws.cell(row=row, column=8).value = a[7]
    ws.cell(row=row, column=9).value = a[8]
    ws.cell(row=row, column=10).value = a[9]
    wb.save(filename)